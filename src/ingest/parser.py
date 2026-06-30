"""Excel 清单解析：智能表头识别 + 仅解析需报价行。"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, List, Optional, Tuple

import pandas as pd

from src.ingest.detector import (
    SheetLayout,
    _row_quantity,
    detect_sheet_layout,
    is_pricing_row,
    should_skip_sheet,
)
from src.ingest.formula_eval import row_quantity_from_ws
from src.ingest.line_types import ParsedLine, ParsedWorkbook, cell_str, to_float
from src.ingest.unit_price_analysis import (
    enrich_from_boq_features,
    parse_unit_price_analysis_sheet,
)
from src.normalize.text import normalize_name, normalize_unit

_cell_str = cell_str
_to_float = to_float


def _get(row: Tuple[Any, ...], mapping: dict, key: str, alt: Optional[str] = None) -> Any:
    idx = mapping.get(alt or key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_sheet(
    df: pd.DataFrame,
    sheet_name: str,
    ws=None,
) -> Tuple[List[ParsedLine], Optional[SheetLayout]]:
    layout = detect_sheet_layout(df, sheet_name)
    if layout is None:
        return [], None

    lines: List[ParsedLine] = []
    section_stack: List[str] = []
    m = layout.column_map
    floor_cols = layout.floor_qty_columns

    for i in range(layout.data_start_row, len(df)):
        row = tuple(df.iloc[i].tolist())
        excel_row = i + 1

        if ws is not None:
            qty_ws = row_quantity_from_ws(ws, excel_row, m, floor_cols)
            pricing_ok = is_pricing_row(row, m, floor_cols, ws=ws, excel_row=excel_row)
        else:
            qty_ws = None
            pricing_ok = is_pricing_row(row, m, floor_cols)

        if not pricing_ok:
            name = _cell_str(_get(row, m, "name"))
            unit = _cell_str(_get(row, m, "unit"))
            if name and not unit:
                section_stack.append(name)
            continue

        def col(key: str, alt: Optional[str] = None) -> Optional[float]:
            return _to_float(_get(row, m, key, alt))

        mm = col("material_main")
        if m.get("_cost_side"):
            mm = _to_float(_get(row, m, "material_main_cost")) or mm
            la = _to_float(_get(row, m, "labor_cost")) or col("labor")
            ma = _to_float(_get(row, m, "material_aux_cost"))
            mc = _to_float(_get(row, m, "machinery_cost"))
            loss = _to_float(_get(row, m, "material_loss_rate_cost")) or col(
                "material_loss_rate"
            )
            cost_up = _to_float(_get(row, m, "cost_unit_price"))
        else:
            la = col("labor")
            ma = col("material_aux")
            mc = col("machinery")
            loss = col("material_loss_rate")
            cost_up = col("cost_unit_price")
        if cost_up is None and any(x is not None for x in (mm, la, ma, mc)):
            cost_up = (mm or 0) * (1 + (loss or 0)) + (la or 0) + (ma or 0) + (mc or 0)
        listed_up = col("unit_price")
        if not m.get("_cost_side"):
            if listed_up is not None and listed_up > 0:
                cost_up = listed_up
            elif (cost_up is None or cost_up == 0) and listed_up:
                cost_up = listed_up

        feat = _cell_str(_get(row, m, "feature"))
        spec = _cell_str(_get(row, m, "spec")) if "spec" in m else ""
        if spec:
            feat = f"{feat}；规格：{spec}" if feat else f"规格：{spec}"
        if sheet_name and feat and not feat.startswith("["):
            feat = f"[{sheet_name}] {feat}"

        qty = qty_ws if qty_ws is not None else _row_quantity(row, m, floor_cols)
        if qty is None and col("unit_price") is not None:
            qty = 1.0
        if floor_cols and qty is not None:
            parts = []
            for idx, label in floor_cols:
                if idx >= len(row):
                    continue
                v = _to_float(row[idx])
                if v is not None:
                    parts.append(f"{label}:{v}")
            if parts:
                extra = "[分层工程量]" + " ".join(parts)
                feat = f"{feat}\n{extra}" if feat else extra
        cost_amt = col("cost_amount")
        if cost_amt is None and cost_up is not None and qty is not None:
            cost_amt = round(cost_up * qty, 2)

        has_cost = cost_up is not None and cost_up > 0
        if not has_cost:
            has_cost = any(
                x is not None and x > 0
                for x in (mm, la, ma, mc, col("management"), col("profit"))
            )

        lines.append(
            ParsedLine(
                sheet_name=sheet_name,
                section_path=" / ".join(section_stack[-4:]),
                seq=_cell_str(_get(row, m, "seq")),
                list_code=_cell_str(_get(row, m, "list_code")),
                name=_cell_str(_get(row, m, "name")),
                feature=feat,
                unit=_cell_str(_get(row, m, "unit")),
                quantity=qty,
                unit_price=col("unit_price"),
                amount=col("amount"),
                remark=_cell_str(_get(row, m, "remark")) if "remark" in m else "",
                material_main=mm,
                material_loss_rate=loss,
                material_aux=ma,
                labor=la,
                machinery=mc,
                management=col("management"),
                profit=col("profit"),
                tax=col("tax"),
                cost_unit_price=cost_up,
                cost_amount=cost_amt,
                row_index=i,
                has_cost_detail=has_cost,
            )
        )
    return lines, layout


def _extract_project_name(xl: pd.ExcelFile, path: Path) -> str:
    for sheet in xl.sheet_names[:4]:
        d = pd.read_excel(xl, sheet_name=sheet, header=None)
        for i in range(min(25, len(d))):
            for j in range(min(8, d.shape[1])):
                v = _cell_str(d.iloc[i, j])
                if v.startswith("工程名称"):
                    rest = _cell_str(d.iloc[i, j + 1]) if j + 1 < d.shape[1] else ""
                    if rest:
                        return rest.replace("：", "").replace(":", "").strip()
                    return v.split("：")[-1].split(":")[-1].strip()
                if "工程名称：" in v or "工程名称:" in v:
                    return re.split(r"工程名称[：:]", v, maxsplit=1)[-1].strip()
    return path.stem


def _is_analysis_sheet(sheet_name: str) -> bool:
    return "综合单价分析" in sheet_name.replace(" ", "")


def parse_workbook(path: str | Path) -> ParsedWorkbook:
    path = Path(path)
    xl = pd.ExcelFile(path)
    project_name = _extract_project_name(xl, path)
    all_lines: List[ParsedLine] = []
    layouts: List[SheetLayout] = []
    analysis_lines: List[ParsedLine] = []

    try:
        from openpyxl import load_workbook

        oxl = load_workbook(path, data_only=False)
    except Exception:
        oxl = None

    for sn in xl.sheet_names:
        if should_skip_sheet(sn) or _is_analysis_sheet(sn):
            continue
        df = pd.read_excel(xl, sheet_name=sn, header=None)
        ws = oxl[sn] if oxl and sn in oxl.sheetnames else None
        parsed, layout = _parse_sheet(df, sn, ws=ws)
        if layout:
            layouts.append(layout)
        all_lines.extend(parsed)

    for sn in xl.sheet_names:
        if not _is_analysis_sheet(sn):
            continue
        df = pd.read_excel(xl, sheet_name=sn, header=None)
        analysis_lines.extend(parse_unit_price_analysis_sheet(df, sn.strip()))

    if analysis_lines:
        enrich_from_boq_features(analysis_lines, all_lines)
        covered = {
            (normalize_name(l.name), normalize_unit(l.unit)) for l in analysis_lines
        }
        for ln in all_lines:
            if _is_analysis_sheet(ln.sheet_name):
                continue
            if (normalize_name(ln.name), normalize_unit(ln.unit)) in covered:
                ln.has_cost_detail = False
        all_lines.extend(analysis_lines)

    cost_lines = sum(1 for ln in all_lines if ln.has_cost_detail)
    if cost_lines > len(all_lines) * 0.3:
        fmt = "internal_cost"
    elif all_lines and cost_lines == 0:
        fmt = "tender"
    else:
        fmt = "mixed"

    return ParsedWorkbook(
        file_path=str(path),
        project_name=project_name or path.stem,
        sheets=list(xl.sheet_names),
        lines=all_lines,
        format_hint=fmt,
        layouts=layouts,
    )
