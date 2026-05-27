"""导出：去重母表 + 全量清单链接公式。"""
from __future__ import annotations

from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.ingest.parser import ParsedLine, parse_workbook
from src.link.dedupe import DedupeItem, build_row_map, dedupe_from_parsed_lines


def export_dedupe_workbook(
    source_path: str,
    output_path: Optional[str | Path] = None,
) -> Tuple[Path, List[DedupeItem]]:
    wb_parse = parse_workbook(source_path)
    items = dedupe_from_parsed_lines(wb_parse.lines)
    src = Path(source_path)
    out = Path(output_path) if output_path else src.parent / f"去重母表_{src.stem}.xlsx"

    rows = []
    for i, it in enumerate(items, start=1):
        rows.append(
            {
                "去重序号": i,
                "项目名称": it.name,
                "项目特征": it.feature,
                "做法摘要": it.method_summary,
                "单位": it.unit,
                "出现次数": it.line_count,
                "工程量合计": round(it.quantity_total, 4),
                "成本单价(待填/已有)": it.cost_unit_price,
                "主材": it.material_main,
                "人工": it.labor,
                "辅材": it.material_aux,
                "机械": it.machinery,
                "综合单价(待填)": None,
            }
        )
    df_master = pd.DataFrame(rows)

    row_map = build_row_map(items)
    map_rows = _build_map_rows(wb_parse.lines, row_map)

    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df_master.to_excel(w, sheet_name="去重母表", index=False)
        pd.DataFrame(map_rows).to_excel(w, sheet_name="行号映射", index=False)
        total = sum(it.line_count for it in items)
        pd.DataFrame(
            [
                {"说明": f"共 {total} 行 → {len(items)} 个唯一项，只需填母表成本"},
                {"说明": "去重键 = 名称 + 单位 + 做法标签（优于 lj.exe 仅名称+单位）"},
                {"说明": "填好后用 link-price 或组价导出（--dedupe-link）生成全量链接"},
            ]
        ).to_excel(w, sheet_name="说明", index=False)

    return out, items


def _build_map_rows(lines: List[ParsedLine], row_map: dict) -> list:
    map_rows = []
    for line in lines:
        if not line.name or not line.unit or line.quantity is None:
            continue
        mr = row_map.get((line.sheet_name, line.row_index))
        map_rows.append(
            {
                "Sheet": line.sheet_name,
                "原行号": line.row_index + 1,
                "序号": line.seq,
                "项目名称": line.name,
                "单位": line.unit,
                "工程量": line.quantity,
                "链接去重序号": mr - 1 if mr else None,
                "链接去重行": mr,
            }
        )
    return map_rows


def export_linked_pricing(
    tender_path: str,
    master_path: str,
    output_path: Optional[str | Path] = None,
) -> Path:
    master = pd.read_excel(master_path, sheet_name="去重母表")
    wb_parse = parse_workbook(tender_path)
    items = dedupe_from_parsed_lines(wb_parse.lines)
    row_map = build_row_map(items)

    src = Path(tender_path)
    out = Path(output_path) if output_path else src.parent / f"链接组价_{src.stem}.xlsx"

    cost_col_name = "成本单价(待填/已有)"
    cost_col = (
        list(master.columns).index(cost_col_name) + 1
        if cost_col_name in master.columns
        else 8
    )
    cost_letter = get_column_letter(cost_col)

    wb = Workbook()
    ws_m = wb.active
    ws_m.title = "去重母表"
    for c, col in enumerate(master.columns, 1):
        ws_m.cell(1, c, col)
    for ri in range(len(master)):
        for c, col in enumerate(master.columns, 1):
            val = master.iloc[ri][col]
            if pd.notna(val):
                ws_m.cell(ri + 2, c, val)

    ws_d = wb.create_sheet("全量清单链接")
    headers = [
        "Sheet",
        "原行",
        "序号",
        "项目名称",
        "特征",
        "单位",
        "工程量",
        "链接成本单价",
        "去重行",
    ]
    for c, h in enumerate(headers, 1):
        ws_d.cell(1, c, h).font = Font(bold=True)

    fill = PatternFill("solid", fgColor="FFF2CC")
    r = 2
    for line in wb_parse.lines:
        if not line.name or not line.unit or line.quantity is None:
            continue
        mr = row_map.get((line.sheet_name, line.row_index))
        ws_d.cell(r, 1, line.sheet_name)
        ws_d.cell(r, 2, line.row_index + 1)
        ws_d.cell(r, 3, line.seq)
        ws_d.cell(r, 4, line.name)
        ws_d.cell(r, 5, (line.feature or "")[:500])
        ws_d.cell(r, 6, line.unit)
        ws_d.cell(r, 7, line.quantity)
        if mr:
            ws_d.cell(r, 8, f"=去重母表!${cost_letter}${mr}")
            ws_d.cell(r, 9, mr)
            for c in range(1, 10):
                ws_d.cell(r, c).fill = fill
        r += 1

    wb.save(out)
    return out
