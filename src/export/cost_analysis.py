"""导出成本分析报告：汇总 + 明细 + 历史候选参考价。"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.db.database import get_connection, resolve_db_path
from src.knowledge.repository import KnowledgeRepository
from src.pricing.engine import PricingEngine
from src.pricing.calc import calc_from_components


def _ref_price(repo: KnowledgeRepository, standard_item_id: int) -> Optional[float]:
    records = repo.get_cost_records_for_item(standard_item_id)
    if not records:
        return None
    agg = repo.aggregate_costs(records)
    price_cfg = repo.settings.get("pricing", {})
    bd = calc_from_components(
        material_main=float(agg.get("material_main") or 0),
        material_loss_rate=float(agg.get("material_loss_rate") or 0),
        labor=float(agg.get("labor") or 0),
        material_aux=float(agg.get("material_aux") or 0),
        machinery=float(agg.get("machinery") or 0),
        management_rate=price_cfg.get("default_management_rate", 0),
        profit_rate=price_cfg.get("default_profit_rate", 0),
        tax_rate=price_cfg.get("default_tax_rate", 0),
    )
    return bd.cost_unit_price


def export_cost_analysis_report(
    job_id: int,
    output_path: Optional[str | Path] = None,
    db_path: Optional[str] = None,
    top_n: int = 3,
) -> Path:
    repo = KnowledgeRepository(db_path)
    engine = PricingEngine(db_path)
    th = engine.thresholds
    conn = get_connection(db_path)
    try:
        job = conn.execute(
            "SELECT j.*, p.name AS project_name FROM pricing_jobs j "
            "JOIN projects p ON p.id=j.project_id WHERE j.id=?",
            (job_id,),
        ).fetchone()
        if not job:
            raise ValueError(f"组价任务不存在: {job_id}")

        lines = conn.execute(
            """SELECT b.*, pl.match_level, pl.confidence, pl.match_note,
                      pl.cost_unit_price AS filled_cost
               FROM pricing_lines pl
               JOIN boq_lines b ON b.id=pl.boq_line_id
               WHERE pl.job_id=?
               ORDER BY b.id""",
            (job_id,),
        ).fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "汇总"

    level_counts: Dict[str, int] = {}
    total_qty_items = 0
    ref_cost_sum = 0.0
    for row in lines:
        lv = row["match_level"] or "D"
        level_counts[lv] = level_counts.get(lv, 0) + 1
        if row["quantity"]:
            total_qty_items += 1

    ws_sum["A1"] = "成本分析报告"
    ws_sum["A1"].font = Font(bold=True, size=16)
    ws_sum["A3"] = "工程名称"
    ws_sum["B3"] = job["project_name"]
    ws_sum["A4"] = "组价任务"
    ws_sum["B4"] = f"job_id={job_id}"
    ws_sum["A5"] = "清单行数"
    ws_sum["B5"] = len(lines)
    ws_sum["A6"] = "有工程量行"
    ws_sum["B6"] = total_qty_items

    ws_sum["A8"] = "匹配等级统计"
    ws_sum["A8"].font = Font(bold=True)
    r = 9
    for lv in ("A", "B", "C", "D"):
        ws_sum[f"A{r}"] = lv + "级"
        ws_sum[f"B{r}"] = level_counts.get(lv, 0)
        r += 1

    ws_sum[f"A{r + 1}"] = "说明"
    ws_sum[f"A{r + 1}"].font = Font(bold=True)
    notes = [
        "A/B级且做法无冲突时系统才自动填入成本单价；C/D级仅列历史参考价，须人工确认。",
        "参考合价 = 工程量 × 候选1历史成本单价中位数（仅供参考，不代表最终报价）。",
        "项目特征含水泥砂浆粘贴/专用勾缝等做法时，与历史「瓷砖胶」项即使名称相近也会降分。",
    ]
    for i, note in enumerate(notes):
        ws_sum[f"A{r + 2 + i}"] = note
        ws_sum.merge_cells(f"A{r + 2 + i}:F{r + 2 + i}")

    ws = wb.create_sheet("成本分析明细")
    headers = [
        "序号",
        "项目名称",
        "项目特征描述",
        "单位",
        "工程量",
        "匹配等级",
        "综合得分",
        "系统说明",
    ]
    for i in range(1, top_n + 1):
        headers.extend(
            [
                f"参考{i}历史项",
                f"参考{i}得分",
                f"参考{i}等级",
                f"参考{i}成本单价",
                f"参考{i}样本数",
            ]
        )
    headers.extend(["参考合价(元)", "参考合价说明"])

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for idx, line in enumerate(lines, 2):
        cands = engine.search(
            line["name"],
            line["feature"] or "",
            line["unit"],
            top_n=top_n,
        )
        ws.cell(row=idx, column=1, value=line["seq"])
        ws.cell(row=idx, column=2, value=line["name"])
        ws.cell(row=idx, column=3, value=line["feature"])
        ws.cell(row=idx, column=4, value=line["unit"])
        ws.cell(row=idx, column=5, value=line["quantity"])
        ws.cell(row=idx, column=6, value=line["match_level"])
        ws.cell(row=idx, column=7, value=round(line["confidence"] or 0, 3))
        ws.cell(row=idx, column=8, value=line["match_note"])

        col = 9
        ref_price = None
        ref_note = ""
        for j, c in enumerate(cands):
            price = _ref_price(repo, c["standard_item_id"])
            if j == 0 and price:
                ref_price = price
            ws.cell(row=idx, column=col, value=c["name"])
            ws.cell(row=idx, column=col + 1, value=round(c["total_score"], 3))
            ws.cell(row=idx, column=col + 2, value=c["level"])
            ws.cell(row=idx, column=col + 3, value=price)
            ws.cell(row=idx, column=col + 4, value=c["samples"])
            if c.get("conflicts"):
                ws.cell(row=idx, column=col).comment = None  # skip comment API
            col += 5

        qty = float(line["quantity"] or 0)
        if ref_price and qty:
            amt = round(ref_price * qty, 2)
            ref_cost_sum += amt
            ws.cell(row=idx, column=col, value=amt)
            if line["match_level"] in ("A", "B"):
                ref_note = "候选与系统匹配一致"
            else:
                ref_note = "仅参考候选1，等级不足需人工核价"
        else:
            ref_note = "无可用历史单价"
        ws.cell(row=idx, column=col + 1, value=ref_note)

    ws_sum["A7"] = "参考合价合计(候选1)"
    ws_sum["B7"] = round(ref_cost_sum, 2)
    ws_sum["C7"] = "须人工审核后采用"

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40

    out_dir = Path(output_path).parent if output_path else resolve_db_path(db_path).parent / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in job["project_name"])[:40]
    file = Path(output_path) if output_path else out_dir / f"成本分析_{safe}_job{job_id}.xlsx"
    wb.save(file)
    return file
