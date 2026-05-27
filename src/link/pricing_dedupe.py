"""组价结果去重：母表填价 + 明细行公式链接（吸收 lj.exe，升级特征/做法键）。"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.link.dedupe import make_dedupe_key


@dataclass
class PricingDedupeItem:
    key: str
    name: str
    feature: str
    unit: str
    method_summary: str
    line_count: int
    boq_line_ids: List[int] = field(default_factory=list)
    material_main: Optional[float] = None
    material_loss_rate: Optional[float] = None
    labor: Optional[float] = None
    material_aux: Optional[float] = None
    machinery: Optional[float] = None
    match_note: str = ""


# 去重母表列 → 明细表 J/K/L/M/N
MASTER_COLS = ("G", "H", "I", "J", "K")
DETAIL_COLS = ("J", "K", "L", "M", "N")
DETAIL_KEYS = (
    "material_main",
    "material_loss_rate",
    "labor",
    "material_aux",
    "machinery",
)


def dedupe_pricing_rows(rows: Sequence[Mapping[str, Any]]) -> Tuple[List[PricingDedupeItem], Dict[int, int]]:
    """pricing_lines + boq 联查结果 → 母项列表、boq_line_id→母表行号(1-based)。"""
    buckets: Dict[str, PricingDedupeItem] = {}
    order: List[str] = []

    for row in rows:
        bid = int(row["boq_line_id"])
        nn, un, sig = make_dedupe_key(row["name"], row["feature"] or "", row["unit"])
        key = f"{nn}|{sig}|{un}"
        if key not in buckets:
            buckets[key] = PricingDedupeItem(
                key=key,
                name=row["name"],
                feature=row["feature"] or "",
                unit=row["unit"],
                method_summary=row.get("method_summary") or "",
                line_count=0,
                match_note=row.get("match_note") or "",
            )
            order.append(key)
        item = buckets[key]
        item.line_count += 1
        item.boq_line_ids.append(bid)
        for k in DETAIL_KEYS:
            if getattr(item, k) is None and row.get(k) is not None:
                setattr(item, k, row[k])
        if not item.match_note and row.get("match_note"):
            item.match_note = row["match_note"]

    items = [buckets[k] for k in order]
    line_to_master: Dict[int, int] = {}
    for i, item in enumerate(items):
        master_row = i + 2  # 第1行表头
        for bid in item.boq_line_ids:
            line_to_master[bid] = master_row
    return items, line_to_master


def write_dedupe_master_sheet(ws, items: List[PricingDedupeItem]) -> None:
    headers = (
        "去重序号",
        "项目名称",
        "项目特征",
        "做法摘要",
        "单位",
        "出现次数",
        "主材",
        "损耗率",
        "人工",
        "辅材",
        "机械",
        "匹配说明",
    )
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)

    for i, it in enumerate(items, start=1):
        r = i + 1
        ws.cell(r, 1, i)
        ws.cell(r, 2, it.name)
        ws.cell(r, 3, (it.feature or "")[:2000])
        ws.cell(r, 4, it.method_summary)
        ws.cell(r, 5, it.unit)
        ws.cell(r, 6, it.line_count)
        for col, val in zip(range(7, 12), DETAIL_KEYS):
            v = getattr(it, val)
            if v is not None:
                ws.cell(r, col, v)
        ws.cell(r, 12, it.match_note)

    ws.freeze_panes = "A2"
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40


def apply_detail_links(
    ws,
    rows: Sequence[Mapping[str, Any]],
    line_to_master: Dict[int, int],
    data_start_row: int,
    boq_id_key: str = "boq_line_id",
) -> int:
    """明细表成本五列改为引用去重母表；返回重复链接行数。"""
    master_counts: Dict[int, int] = {}
    for mr in line_to_master.values():
        master_counts[mr] = master_counts.get(mr, 0) + 1

    linked = 0
    fill = PatternFill("solid", fgColor="FFF2CC")
    for i, row in enumerate(rows):
        r = data_start_row + i
        bid = int(row[boq_id_key])
        mr = line_to_master.get(bid)
        if not mr:
            continue
        for dcol, mcol in zip(DETAIL_COLS, MASTER_COLS):
            ws[f"{dcol}{r}"] = f"=去重母表!${mcol}${mr}"
        if master_counts.get(mr, 0) > 1:
            for c in range(1, 19):
                ws.cell(r, c).fill = fill
            linked += 1
    return linked


def dedupe_stats(items: List[PricingDedupeItem], total_lines: int) -> Dict[str, Any]:
    unique = len(items)
    saved = total_lines - unique
    pct = round(100 * saved / total_lines, 1) if total_lines else 0
    return {
        "total_lines": total_lines,
        "unique_items": unique,
        "duplicate_lines": saved,
        "save_percent": pct,
    }
