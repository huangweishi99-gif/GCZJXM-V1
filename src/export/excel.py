"""导出投标方单价分析表（18列，含 Excel 公式，费率默认 0）。"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from src.db.database import get_connection, resolve_db_path
from src.link.pricing_dedupe import (
    apply_detail_links,
    dedupe_pricing_rows,
    dedupe_stats,
    write_dedupe_master_sheet,
)


def _load_template() -> dict:
    p = Path(__file__).resolve().parents[2] / "config" / "column_templates.json"
    return json.loads(p.read_text(encoding="utf-8"))["bidder_export"]


def _fetch_job_rows(conn, job_id: int) -> list:
    return conn.execute(
        """SELECT b.id AS boq_line_id, b.seq, b.name, b.feature, b.unit, b.quantity, b.remark,
                  b.method_summary,
                  pl.match_level, pl.match_note,
                  pl.material_main, pl.material_loss_rate, pl.material_aux,
                  pl.labor, pl.machinery, pl.management, pl.profit, pl.tax,
                  pl.cost_unit_price, pl.cost_amount, pl.unit_price
           FROM pricing_lines pl
           JOIN boq_lines b ON b.id = pl.boq_line_id
           WHERE pl.job_id=?
           ORDER BY b.id""",
        (job_id,),
    ).fetchall()


def _write_detail_sheet(ws, rows, tpl: dict, job: dict) -> None:
    rates = tpl["default_rates"]
    ws.merge_cells("A1:R1")
    ws["A1"] = "E.1 分部分项工程项目清单计价表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:B2")
    ws["A2"] = f"工程名称：{job['project_name']}"

    ws["A3"], ws["B3"], ws["C3"], ws["D3"], ws["E3"] = (
        "序号",
        "项目名称",
        "项目特征描述",
        "计量\n单位",
        "工程量",
    )
    ws.merge_cells("F3:H3")
    ws["F3"] = "金额(元)"
    ws.merge_cells("I3:N3")
    ws["I3"] = "成本单价分析"
    ws["O3"], ws["P3"], ws["Q3"], ws["R3"] = "合价", "管理费", "利润", "税金"

    ws.merge_cells("A3:A4")
    ws.merge_cells("B3:B4")
    ws.merge_cells("C3:C4")
    ws.merge_cells("D3:D4")
    ws.merge_cells("E3:E4")

    ws["F4"], ws["G4"], ws["H4"] = "综合单价", "合价", "备注"
    ws["I4"], ws["J4"], ws["K4"], ws["L4"], ws["M4"], ws["N4"] = (
        "成本单价",
        "主材",
        "损耗率",
        "人工",
        "辅材",
        "机械",
    )
    ws["O4"] = "合价"
    ws["P4"] = rates["management_rate"]
    ws["Q4"] = rates["profit_rate"]
    ws["R4"] = rates["tax_rate"]

    start = tpl["data_start_row"]
    f = tpl["formulas"]
    for i, row in enumerate(rows):
        r = start + i
        ws[f"A{r}"] = row["seq"]
        ws[f"B{r}"] = row["name"]
        ws[f"C{r}"] = row["feature"]
        ws[f"D{r}"] = row["unit"]
        ws[f"E{r}"] = row["quantity"]
        if row["remark"]:
            ws[f"H{r}"] = row["remark"]
        elif row["match_note"]:
            ws[f"H{r}"] = row["match_note"]

        for col_letter, key in (
            ("J", "material_main"),
            ("K", "material_loss_rate"),
            ("L", "labor"),
            ("M", "material_aux"),
            ("N", "machinery"),
        ):
            if row[key] is not None:
                ws[f"{col_letter}{r}"] = row[key]

        ws[f"I{r}"] = f["cost_unit_price"].format(row=r)
        ws[f"O{r}"] = f["cost_amount"].format(row=r)
        ws[f"P{r}"] = f["management"].format(row=r)
        ws[f"Q{r}"] = f["profit"].format(row=r)
        ws[f"R{r}"] = f["tax"].format(row=r)
        ws[f"F{r}"] = f["unit_price"].format(row=r)
        ws[f"G{r}"] = f["amount"].format(row=r)


def export_pricing_job(
    job_id: int,
    output_path: Optional[str | Path] = None,
    db_path: Optional[str] = None,
    use_dedupe_link: bool = False,
) -> Path:
    tpl = _load_template()
    conn = get_connection(db_path)
    try:
        job = conn.execute(
            "SELECT j.*, p.name AS project_name FROM pricing_jobs j "
            "JOIN projects p ON p.id=j.project_id WHERE j.id=?",
            (job_id,),
        ).fetchone()
        if not job:
            raise ValueError(f"组价任务不存在: {job_id}")

        raw_rows = _fetch_job_rows(conn, job_id)
        rows = [dict(r) for r in raw_rows]

        out = Path(output_path) if output_path else resolve_db_path(db_path).parent / "exports"
        out.mkdir(parents=True, exist_ok=True)
        safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in job["project_name"])[:40]
        suffix = "_去重链接" if use_dedupe_link else ""
        file = out / f"投标方单价分析_{safe_name}_job{job_id}{suffix}.xlsx"

        wb = Workbook()
        dedupe_meta: Optional[Dict[str, Any]] = None

        if use_dedupe_link and rows:
            items, line_to_master = dedupe_pricing_rows(rows)
            dedupe_meta = dedupe_stats(items, len(rows))

            ws_master = wb.active
            ws_master.title = "去重母表"
            write_dedupe_master_sheet(ws_master, items)

            ws = wb.create_sheet("单价分析")
            _write_detail_sheet(ws, rows, tpl, job)
            linked = apply_detail_links(ws, rows, line_to_master, tpl["data_start_row"])
            dedupe_meta["linked_duplicate_rows"] = linked
        else:
            ws = wb.active
            ws.title = "单价分析"
            _write_detail_sheet(ws, rows, tpl, job)

        meta = wb.create_sheet("说明")
        meta.append(["字段", "说明"])
        meta.append(["报价行", "有项目名称、单位、工程量的行"])
        meta.append(["费率", "管理费/利润/税金 当前固定 0%"])
        if use_dedupe_link and dedupe_meta:
            meta.append(["去重链接", "黄色行=重复项，成本五列引用「去重母表」，改母表全表更新"])
            meta.append(
                [
                    "去重统计",
                    f"{dedupe_meta['total_lines']}行 → {dedupe_meta['unique_items']}项，"
                    f"少填 {dedupe_meta['duplicate_lines']} 行 ({dedupe_meta['save_percent']}%)",
                ]
            )
        meta.append(["公式", "见 docs/表头对照与计算公式.md"])

        wb.save(file)
        conn.execute(
            "UPDATE pricing_jobs SET output_file=?, status='exported' WHERE id=?",
            (str(file), job_id),
        )
        conn.commit()
        return file
    finally:
        conn.close()
