"""在原甲方招标 Excel 上填入投标方成本列（18列结构）并做同表去重链接。"""
from __future__ import annotations

import json
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.db.database import get_connection, resolve_db_path
from src.ingest.detector import SheetLayout, detect_sheet_layout, MAX_REASONABLE_COL
from src.knowledge.query import PricingContext
from src.knowledge.repository import KnowledgeRepository
from src.link.dedupe import make_dedupe_key
from src.pricing.calc import calc_from_components
from src.pricing.engine import PricingEngine
from src.pricing.component_judge import judge_line_components, judge_to_reference_tuple

# 追加的成本区字段（在已有列之后）
COST_FIELDS = (
    "cost_unit_price",
    "material_main",
    "material_loss_rate",
    "labor",
    "material_aux",
    "machinery",
    "cost_amount",
    "management",
    "profit",
    "tax",
)

COST_HEADERS = (
    "成本单价",
    "主材",
    "损耗率",
    "人工",
    "辅材",
    "机械",
    "合价",
    "管理费",
    "利润",
    "税金",
)

MASTER_FILL = PatternFill("solid", fgColor="C6EFCE")
SLAVE_FILL = PatternFill("solid", fgColor="FFF2CC")

# 甲方报价表（综合单价=其中合计）：右侧成本区 + G/H/I 桥接，不改 F/J 原公式逻辑
TENDER_BREAKDOWN_COST_HEADERS = (
    (12, None, "复核工程量"),
    (13, "cost_unit_price", "成本价"),
    (14, "material_main", "主材费"),
    (15, "material_loss_rate", "主材损耗率"),
    (16, "labor", "人工费"),
    (17, "material_aux", "辅材费"),
    (18, "machinery", "机械费"),
    (19, "cost_amount", "成本合计"),
)


@dataclass
class SheetColumns:
    """Sheet 内 1-based 列号。"""
    seq: Optional[int] = None
    list_code: Optional[int] = None
    name: Optional[int] = None
    feature: Optional[int] = None
    unit: Optional[int] = None
    quantity: Optional[int] = None
    unit_price: Optional[int] = None
    amount: Optional[int] = None
    remark: Optional[int] = None
    cost_unit_price: Optional[int] = None
    material_main: Optional[int] = None
    material_loss_rate: Optional[int] = None
    labor: Optional[int] = None
    material_aux: Optional[int] = None
    machinery: Optional[int] = None
    cost_amount: Optional[int] = None
    management: Optional[int] = None
    profit: Optional[int] = None
    tax: Optional[int] = None
    # 其中：人工费/材料费/其他费（开市客等甲方报价表）
    tender_labor: Optional[int] = None
    tender_material: Optional[int] = None
    tender_other: Optional[int] = None
    preserve_tender_breakdown: bool = False

    def col(self, field: str) -> Optional[int]:
        return getattr(self, field, None)

    def letter(self, field: str) -> Optional[str]:
        c = self.col(field)
        return get_column_letter(c) if c else None


def _load_tpl() -> dict:
    p = Path(__file__).resolve().parents[2] / "config" / "column_templates.json"
    return json.loads(p.read_text(encoding="utf-8"))["bidder_export"]


def _is_tender_breakdown_sheet(ws, layout: SheetLayout) -> bool:
    """识别「综合单价 + 其中(人工/材料/其他)」的甲方报价表（如开市客）。"""
    candidates = []
    if layout.subheader_row is not None:
        candidates.append(layout.subheader_row + 1)
    candidates.extend([layout.header_row + 2, layout.header_row + 3])
    for sr in candidates:
        labels = []
        for c in (7, 8, 9):
            v = ws.cell(sr, c).value
            if v:
                labels.append(str(v).replace(" ", ""))
        if not labels:
            continue
        text = "".join(labels)
        if "人工费" in text and "材料费" in text and ("其他费" in text or "机械" in text):
            return True
    return False


def _apply_tender_breakdown_columns(ws, layout: SheetLayout, cols: SheetColumns) -> None:
    """固定右侧成本列位，保留左侧招标综合单价列不动。"""
    cols.preserve_tender_breakdown = True
    cols.tender_labor = 7
    cols.tender_material = 8
    cols.tender_other = 9
    for col_idx, field, _title in TENDER_BREAKDOWN_COST_HEADERS:
        if field:
            setattr(cols, field, col_idx)
    # 费率系数行（与用户校正表一致）
    ws.cell(1, 13, 0)
    ws.cell(1, 14, 1)
    ws.cell(1, 16, 1)
    ws.cell(1, 17, 1)


def _ensure_tender_breakdown_headers(ws, layout: SheetLayout, cols: SheetColumns) -> None:
    hr = layout.header_row + 1
    sr = layout.subheader_row + 1 if layout.subheader_row is not None else layout.header_row + 3
    for col_idx, _field, title in TENDER_BREAKDOWN_COST_HEADERS:
        if title:
            ws.cell(hr, col_idx, title)
            ws.cell(hr, col_idx).font = Font(bold=True)
            ws.cell(hr, col_idx).alignment = Alignment(horizontal="center", wrap_text=True)
            if col_idx >= 13:
                ws.cell(sr, col_idx, title)
                ws.cell(sr, col_idx).font = Font(bold=True)


def _effective_last_column(ws, layout: SheetLayout) -> int:
    """按表头实际占用列计算末列，避免 ws.max_column 虚高导致成本区写到 XFD。"""
    last = 0
    rows = [layout.header_row + 1]
    if layout.subheader_row is not None:
        rows.append(layout.subheader_row + 1)
    elif layout.header_row + 2 <= ws.max_row:
        rows.append(layout.header_row + 2)
    for r in rows:
        for c in range(1, MAX_REASONABLE_COL + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                last = max(last, c)
    # 清单询等：数据行在「检查」列(≈13)有占位，成本区须在其后
    for r in range(layout.data_start_row + 1, min(layout.data_start_row + 40, ws.max_row + 1)):
        for c in range(12, min(20, MAX_REASONABLE_COL + 1)):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() not in ("", "None"):
                last = max(last, c)
    return last


def _layout_to_columns(layout: SheetLayout, max_col: int, ws=None) -> SheetColumns:
    m = layout.column_map
    cols = SheetColumns()
    # 双表头（招标列 + 成本分析列）时优先用 *_cost 侧列号
    cost_field_alias = {
        "material_main": "material_main_cost",
        "labor": "labor_cost",
        "material_aux": "material_aux_cost",
        "machinery": "machinery_cost",
        "material_loss_rate": "material_loss_rate_cost",
    }
    use_cost_side = bool(m.get("_cost_side"))

    for f in (
        "seq",
        "list_code",
        "name",
        "feature",
        "unit",
        "quantity",
        "unit_price",
        "amount",
        "remark",
        "cost_unit_price",
        "material_main",
        "material_loss_rate",
        "labor",
        "material_aux",
        "machinery",
        "cost_amount",
        "management",
        "profit",
        "tax",
    ):
        key = f
        if use_cost_side:
            alias = cost_field_alias.get(f)
            if alias and alias in m:
                key = alias
        if key in m:
            setattr(cols, f, m[key] + 1)
        elif f in m:
            setattr(cols, f, m[f] + 1)

    if cols.amount is None and cols.unit_price is not None:
        cols.amount = cols.unit_price + 1

    # 地库等双表：右侧「合计」列即成本合价
    if use_cost_side and cols.cost_amount is None and cols.amount is not None:
        cols.cost_amount = cols.amount

    # 开市客等甲方报价表：固定成本列位，不追加投标方18列
    if ws is not None and _is_tender_breakdown_sheet(ws, layout):
        _apply_tender_breakdown_columns(ws, layout, cols)
        return cols

    # 已有成本区（投标方18列等）；虚高 max_column 时按表头末列+1 追加
    need_append = cols.cost_unit_price is None or cols.cost_unit_price > MAX_REASONABLE_COL
    if need_append:
        if ws is not None:
            start = _effective_last_column(ws, layout) + 1
        else:
            start = min(max_col, MAX_REASONABLE_COL) + 1
        for i, f in enumerate(COST_FIELDS):
            setattr(cols, f, start + i)

    return cols


def _ensure_cost_headers(ws, layout: SheetLayout, cols: SheetColumns) -> None:
    """在表头行写入成本区列名（不改变原有列）。"""
    hr = layout.header_row + 1
    sr = (layout.subheader_row + 1) if layout.subheader_row is not None else hr + 1
    c0 = cols.cost_unit_price
    c1 = cols.tax
    if not c0 or not c1:
        return

    ws.cell(hr, c0, "成本单价分析")
    ws.cell(hr, c0).font = Font(bold=True)
    ws.cell(hr, c0).alignment = Alignment(horizontal="center")
    if c1 > c0:
        ws.merge_cells(start_row=hr, start_column=c0, end_row=hr, end_column=c1)

    for field, title in zip(COST_FIELDS, COST_HEADERS):
        c = cols.col(field)
        if c:
            ws.cell(sr, c, title)
            ws.cell(sr, c).font = Font(bold=True)
            ws.cell(sr, c).alignment = Alignment(horizontal="center", wrap_text=True)

    rates = _load_tpl()["default_rates"]
    for field, rate in (
        ("management", rates["management_rate"]),
        ("profit", rates["profit_rate"]),
        ("tax", rates["tax_rate"]),
    ):
        c = cols.col(field)
        if c:
            ws.cell(sr, c, rate)


def _resolve_costs(
    row: Mapping[str, Any],
    engine: PricingEngine,
    repo: KnowledgeRepository,
    reference_fill: bool,
    ctx: Optional[PricingContext] = None,
) -> Tuple[Optional[dict], str]:
    """整项须名称+特征相似；否则按材料规格查主材价。"""
    has = any(row.get(k) is not None for k in COST_FIELDS[:6])
    existing = None
    if has:
        existing = {
            "material_main": row.get("material_main") or 0,
            "material_loss_rate": row.get("material_loss_rate") or 0,
            "labor": row.get("labor") or 0,
            "material_aux": row.get("material_aux") or 0,
            "machinery": row.get("machinery") or 0,
        }
    if existing and any(existing.get(k) for k in existing):
        return existing, row.get("match_note") or ""
    judgment = judge_line_components(
        row["name"],
        row["feature"] or "",
        row["unit"],
        engine,
        repo,
        reference_fill=reference_fill,
        ctx=ctx,
    )
    return judge_to_reference_tuple(judgment)


def _write_master_row(
    ws,
    excel_row: int,
    cols: SheetColumns,
    components: dict,
    qty: float,
    remark: str,
) -> None:
    """母行：写入人材机数值 + 公式链。"""
    rates = _load_tpl()["default_rates"]
    bd = calc_from_components(
        material_main=float(components.get("material_main") or 0),
        material_loss_rate=float(components.get("material_loss_rate") or 0),
        labor=float(components.get("labor") or 0),
        material_aux=float(components.get("material_aux") or 0),
        machinery=float(components.get("machinery") or 0),
        management_rate=rates["management_rate"],
        profit_rate=rates["profit_rate"],
        tax_rate=rates["tax_rate"],
    )

    main_l = cols.letter("material_main")
    loss_l = cols.letter("material_loss_rate")
    labor_l = cols.letter("labor")
    aux_l = cols.letter("material_aux")
    mach_l = cols.letter("machinery")
    cost_l = cols.letter("cost_unit_price")
    q_col = cols.letter("quantity")
    cost_amt_l = cols.letter("cost_amount")
    mgmt_l = cols.letter("management")
    profit_l = cols.letter("profit")
    tax_l = cols.letter("tax")
    f_up = cols.letter("unit_price")
    g_amt = cols.letter("amount")
    sr = excel_row

    if cols.col("material_main"):
        ws.cell(sr, cols.col("material_main"), bd.material_main or None)
    if cols.col("material_loss_rate"):
        # 须写 0 而非留空，否则招标表原格残留（如特征序号「4.」）会被公式当损耗率
        ws.cell(sr, cols.col("material_loss_rate"), bd.material_loss_rate)
    if cols.col("labor"):
        ws.cell(sr, cols.col("labor"), bd.labor or None)
    if cols.col("material_aux"):
        ws.cell(sr, cols.col("material_aux"), bd.material_aux or None)
    if cols.col("machinery"):
        ws.cell(sr, cols.col("machinery"), bd.machinery or None)

    if cols.preserve_tender_breakdown:
        _write_tender_breakdown_bridge(ws, sr, cols)
        if remark and cols.col("remark"):
            existing = ws.cell(sr, cols.col("remark")).value
            if not existing:
                ws.cell(sr, cols.col("remark"), remark[:500])
        for field in (
            "cost_unit_price",
            "material_main",
            "material_loss_rate",
            "labor",
            "material_aux",
            "machinery",
            "cost_amount",
        ):
            c = cols.col(field)
            if c:
                ws.cell(sr, c).fill = MASTER_FILL
        return

    if main_l and loss_l and labor_l and aux_l and mach_l and cost_l:
        ws[f"{cost_l}{sr}"] = (
            f"={main_l}{sr}+{main_l}{sr}*{loss_l}{sr}"
            f"+{labor_l}{sr}+{aux_l}{sr}+{mach_l}{sr}"
        )

    if cost_l and q_col and cost_amt_l:
        ws[f"{cost_amt_l}{sr}"] = f"={cost_l}{sr}*{q_col}{sr}"

    if main_l and mgmt_l and profit_l and tax_l:
        base = (
            f"({main_l}{sr}+{main_l}{sr}*{loss_l}{sr}"
            f"+{labor_l}{sr}+{aux_l}{sr}+{mach_l}{sr})"
        )
        ws[f"{mgmt_l}{sr}"] = f"={base}*${mgmt_l}$4"
        ws[f"{profit_l}{sr}"] = f"=({base}+{mgmt_l}{sr})*${profit_l}$4"
        ws[f"{tax_l}{sr}"] = f"=({base}+{mgmt_l}{sr}+{profit_l}{sr})*${tax_l}$4"

    if cost_l and f_up and mgmt_l and profit_l and tax_l:
        ws[f"{f_up}{sr}"] = f"={cost_l}{sr}+{mgmt_l}{sr}+{profit_l}{sr}+{tax_l}{sr}"

    if f_up and g_amt and q_col and (not cost_amt_l or g_amt != cost_amt_l):
        ws[f"{g_amt}{sr}"] = f"={q_col}{sr}*{f_up}{sr}"

    if remark and cols.col("remark"):
        existing = ws.cell(sr, cols.col("remark")).value
        if not existing:
            ws.cell(sr, cols.col("remark"), remark[:500])

    for field in COST_FIELDS + ("unit_price", "amount"):
        c = cols.col(field)
        if c:
            ws.cell(sr, c).fill = MASTER_FILL


def _write_tender_breakdown_bridge(ws, excel_row: int, cols: SheetColumns) -> None:
    """甲方报价表：成本区填数 + 其中(G/H/I)桥接到成本列；综合单价 F=G+H+I 不改投标公式。"""
    sr = excel_row
    n_l = cols.letter("material_main")
    o_l = cols.letter("material_loss_rate")
    p_l = cols.letter("labor")
    q_l = cols.letter("material_aux")
    r_l = cols.letter("machinery")
    m_l = cols.letter("cost_unit_price")
    s_l = cols.letter("cost_amount")
    e_l = cols.letter("quantity")
    f_l = cols.letter("unit_price")
    g_l = get_column_letter(cols.tender_labor or 7)
    h_l = get_column_letter(cols.tender_material or 8)
    i_l = get_column_letter(cols.tender_other or 9)
    j_l = cols.letter("amount")

    if n_l and o_l and p_l and q_l and r_l and m_l:
        ws[f"{m_l}{sr}"] = (
            f"={n_l}{sr}+{n_l}{sr}*{o_l}{sr}+{p_l}{sr}+{q_l}{sr}+{r_l}{sr}"
        )
    if m_l and e_l and s_l:
        ws[f"{s_l}{sr}"] = f"={e_l}{sr}*{m_l}{sr}"
    if p_l and g_l:
        ws[f"{g_l}{sr}"] = f"={p_l}{sr}*$P$1"
    if n_l and o_l and q_l and r_l and h_l:
        ws[f"{h_l}{sr}"] = (
            f"=({n_l}{sr}+{n_l}{sr}*{o_l}{sr})*$N$1+{q_l}{sr}*$Q$1+{r_l}{sr}"
        )
    if m_l and i_l:
        ws[f"{i_l}{sr}"] = f"={m_l}{sr}*$M$1"
  # 仅恢复/保持招标综合单价与合价公式，不写入成本投标价
    if f_l and g_l and h_l and i_l:
        ws[f"{f_l}{sr}"] = f"={g_l}{sr}+{h_l}{sr}+{i_l}{sr}"
    if j_l and f_l and e_l:
        ws[f"{j_l}{sr}"] = f"={f_l}{sr}*{e_l}{sr}"


def _link_slave_row(ws, excel_row: int, master_row: int, cols: SheetColumns) -> None:
    """重复行：同表链接到母行（lj 思路）。"""
    if cols.preserve_tender_breakdown:
        _link_slave_row_tender_breakdown(ws, excel_row, master_row, cols)
        return

    link_fields = (
        "cost_unit_price",
        "material_main",
        "material_loss_rate",
        "labor",
        "material_aux",
        "machinery",
        "cost_amount",
        "management",
        "profit",
        "tax",
        "unit_price",
        "amount",
    )
    fill = SLAVE_FILL
    for field in link_fields:
        c = cols.col(field)
        m = cols.letter(field)
        if c and m:
            ws.cell(excel_row, c, f"={m}${master_row}")
            ws.cell(excel_row, c).fill = fill


def _link_slave_row_tender_breakdown(
    ws, excel_row: int, master_row: int, cols: SheetColumns
) -> None:
    """甲方报价表重复行：成本区绝对链接母行，其中/综合单价/合价保持行内桥接公式。"""
    sr = excel_row
    for field in (
        "material_main",
        "material_loss_rate",
        "labor",
        "material_aux",
        "machinery",
        "cost_unit_price",
        "cost_amount",
    ):
        c = cols.col(field)
        m = cols.letter(field)
        if c and m:
            ws.cell(sr, c, f"={m}${master_row}")
            ws.cell(sr, c).fill = SLAVE_FILL
    _write_tender_breakdown_bridge(ws, sr, cols)
    for c in (cols.tender_labor, cols.tender_material, cols.tender_other, cols.unit_price, cols.amount):
        if c:
            ws.cell(sr, c).fill = SLAVE_FILL


def fill_tender_inplace(
    tender_path: str | Path,
    job_id: int,
    output_path: Optional[str | Path] = None,
    db_path: Optional[str] = None,
    reference_fill: bool = True,
) -> Path:
    """
    复制原招标清单，保留原表结构，追加/识别成本列，填入组价并同表去重链接。
    """
    src = Path(tender_path)
    if output_path:
        out = Path(output_path)
    else:
        out = resolve_db_path(db_path).parent / "exports" / f"{src.stem}_组价.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, out)

    conn = get_connection(db_path)
    repo = KnowledgeRepository(db_path)
    engine = PricingEngine(db_path)
    try:
        rows = conn.execute(
            """SELECT b.id AS boq_line_id, b.sheet_name, b.row_index, b.seq, b.name, b.feature,
                      b.unit, b.quantity, b.remark,
                      pl.match_level, pl.match_note,
                      pl.material_main, pl.material_loss_rate, pl.material_aux,
                      pl.labor, pl.machinery, pl.cost_unit_price
               FROM pricing_lines pl
               JOIN boq_lines b ON b.id = pl.boq_line_id
               WHERE pl.job_id=?
               ORDER BY b.sheet_name, b.row_index""",
            (job_id,),
        ).fetchall()
        job = conn.execute(
            "SELECT j.*, p.name AS project_name FROM pricing_jobs j "
            "JOIN projects p ON p.id=j.project_id WHERE j.id=?",
            (job_id,),
        ).fetchone()
    finally:
        conn.close()

    if not job:
        raise ValueError(f"组价任务不存在: {job_id}")

    ctx = repo.get_project_context(int(job["project_id"]))

    wb = load_workbook(out)
    by_sheet: Dict[str, List[dict]] = {}
    for r in rows:
        by_sheet.setdefault(r["sheet_name"], []).append(dict(r))

    linked = 0
    filled = 0
    used_tender_breakdown = False

    for sheet_name, sheet_rows in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        df = pd.read_excel(out, sheet_name=sheet_name, header=None)
        layout = detect_sheet_layout(df, sheet_name)
        if not layout:
            continue

        cols = _layout_to_columns(layout, ws.max_column, ws=ws)
        if cols.preserve_tender_breakdown:
            used_tender_breakdown = True
            _ensure_tender_breakdown_headers(ws, layout, cols)
        else:
            _ensure_cost_headers(ws, layout, cols)

        masters: Dict[str, int] = {}
        for row in sheet_rows:
            excel_row = int(row["row_index"]) + 1
            nn, un, sig = make_dedupe_key(row["name"], row["feature"] or "", row["unit"])
            key = f"{nn}|{sig}|{un}"

            if key not in masters:
                components, note = _resolve_costs(
                    row, engine, repo, reference_fill, ctx=ctx
                )
                if components:
                    qty = float(row["quantity"] or 0)
                    _write_master_row(ws, excel_row, cols, components, qty, note)
                    masters[key] = excel_row
                    filled += 1
                else:
                    if note and cols.col("remark"):
                        ws.cell(excel_row, cols.col("remark"), note[:500])
            else:
                _link_slave_row(ws, excel_row, masters[key], cols)
                linked += 1

    meta = wb.create_sheet("_组价说明")
    meta.append(["工程", job["project_name"]])
    meta.append(["组价任务", job_id])
    meta.append(["母行填价", filled])
    meta.append(["链接重复行", linked])
    note = (
        "甲方报价表：综合单价=其中合计；成本区填数后 G/H/I 桥接；重复项链接母行"
        if used_tender_breakdown
        else "保留原表；成本区按投标方需求表头填价；相同项链接到首次出现行"
    )
    meta.append(["说明", note])

    wb.save(out)
    return out
