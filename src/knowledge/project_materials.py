# -*- coding: utf-8 -*-
"""项目主材编号价表（如售楼处主材料.xlsx）导入与查询。"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl

from src.db.database import get_connection
from src.knowledge.facts import upsert_material_fact_from_line
from src.normalize.text import normalize_name, normalize_unit

MATERIAL_CODE_RE = re.compile(
    r"([A-Z]{2}-\d{1,2}(?:\.\d{1,2})?[a-zA-Z]?)(?![A-Z0-9.\-])",
    re.I,
)

_CODE_CATEGORY = {
    "PT": "涂料",
    "GL": "玻璃",
    "MT": "金属",
    "AL": "金属",
    "CA": "地毯",
    "UP": "布艺皮革",
    "ST": "石材",
    "CT": "瓷砖",
    "CP": "毛毡",
    "WD": "木作",
    "SP": "其他饰面",
    "SF": "软装",
    "MR": "镜子",
    "DF": "其他饰面",
}


_SUPPLEMENTS_PATH = Path(__file__).resolve().parents[2] / "config" / "project_material_supplements.json"
_supplements_cache: Optional[dict] = None


def _load_supplements() -> dict:
    global _supplements_cache
    if _supplements_cache is not None:
        return _supplements_cache
    if not _SUPPLEMENTS_PATH.exists():
        _supplements_cache = {}
        return _supplements_cache
    _supplements_cache = json.loads(_SUPPLEMENTS_PATH.read_text(encoding="utf-8"))
    return _supplements_cache


def _supplement_material_row(
    code: str,
    *,
    project_ref: Optional[str],
    unit: str,
    city: str = "",
) -> Optional[dict]:
    refs = []
    if project_ref:
        refs.append(project_ref)
    if city:
        refs.append(f"city:{city}")
    for ref in refs:
        sup = _load_supplements().get(ref, {}).get(code.upper())
        if not sup:
            continue
        un = normalize_unit(str(sup.get("unit_norm") or "㎡"))
        line_u = normalize_unit(unit)
        if line_u and un and line_u != un and line_u not in ("m2", "㎡") and un in ("m2", "㎡"):
            pass
        elif line_u and un and line_u != un:
            continue
        return {
            "material_code": code.upper(),
            "material_name": str(sup.get("material_name") or code),
            "material_main": float(sup["material_main"]),
            "unit_norm": un,
            "project_name": ref,
            "project_ref": ref,
        }
    return None


def _material_try_codes(code: str) -> List[str]:
    """编号查价顺序：WD-03→WD-04；ST-01.02→ST-01。"""
    try_codes = [code.upper()]
    if code.upper() == "WD-03":
        try_codes.append("WD-04")
    m_sub = re.match(r"^(ST-\d+)\.\d+", code, re.I)
    if m_sub:
        parent = m_sub.group(1).upper()
        if parent not in try_codes:
            try_codes.append(parent)
    return try_codes


def extract_material_codes(name: str, feature: str = "") -> List[str]:
    """从名称+特征提取物料编号（PT-01、CT-01.02、ST-04a 等），去重保序（长编号优先）。"""
    text = f"{name or ''} {feature or ''}"
    seen: set[str] = set()
    raw: List[str] = []
    for m in MATERIAL_CODE_RE.finditer(text):
        code = m.group(1).upper()
        if code not in seen:
            seen.add(code)
            raw.append(code)
    raw.sort(key=lambda c: (-len(c), text.upper().find(c)))
    return raw


def _category_from_code(code: str) -> str:
    prefix = code.split("-")[0].upper()
    return _CODE_CATEGORY.get(prefix, "饰面")


def import_project_materials_xlsx(
    file_path: str | Path,
    *,
    project_name: str,
    project_ref: str,
    city: str = "",
    price_tier: str = "mid",
    db_path: Optional[str] = None,
) -> dict:
    """
    导入项目主材表（物料编号、名称、单位、不含税主材价）。
    写入 project_material_prices，并同步 material_price_facts（key=code:PT-01）。
    """
    fp = Path(file_path)
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    rows_in: list = []
    section = ""

    for r in range(1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        spec = ws.cell(r, 3).value
        area = ws.cell(r, 4).value
        unit = ws.cell(r, 6).value
        price = ws.cell(r, 7).value
        remark = ws.cell(r, 8).value

        if code and isinstance(code, str) and not re.match(r"^[A-Z]{2}-\d", code.strip(), re.I):
            if name and str(name).strip() and not price:
                section = str(name).strip()
            continue

        if not code or not isinstance(code, str):
            continue
        code = code.strip().upper()
        if not re.match(r"^[A-Z]{2}-\d", code):
            continue
        if price is None:
            continue
        try:
            price_f = float(price)
        except (TypeError, ValueError):
            continue
        if price_f <= 0:
            continue

        mat_name = str(name).strip() if name else code
        un = normalize_unit(str(unit) if unit else "㎡")
        rows_in.append(
            {
                "material_code": code,
                "material_name": mat_name,
                "name_norm": normalize_name(mat_name),
                "spec_text": str(spec).strip() if spec else "",
                "use_area": str(area).strip() if area else "",
                "unit_norm": un,
                "material_main": price_f,
                "category": section or _category_from_code(code),
                "remark": str(remark).strip() if remark else "",
            }
        )

    conn = get_connection(db_path)
    try:
        conn.execute(
            "DELETE FROM project_material_prices WHERE project_ref=? AND source_file=?",
            (project_ref, str(fp)),
        )
        for row in rows_in:
            conn.execute(
                """INSERT OR REPLACE INTO project_material_prices
                   (material_code, material_name, name_norm, spec_text, use_area,
                    unit_norm, material_main, category, project_name, project_ref,
                    city, price_tier, source_file, remark)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    row["material_code"],
                    row["material_name"],
                    row["name_norm"],
                    row["spec_text"],
                    row["use_area"],
                    row["unit_norm"],
                    row["material_main"],
                    row["category"],
                    project_name,
                    project_ref,
                    city or "",
                    price_tier or "mid",
                    str(fp),
                    row["remark"],
                ),
            )
            conn.execute(
                """INSERT INTO material_price_facts
                   (material_key, material_category, spec_text, unit_norm, city, price_tier,
                    material_main, material_loss_rate, sample_count)
                   VALUES (?,?,?,?,?,?,?,?,1)
                   ON CONFLICT(material_key, unit_norm, city, price_tier) DO UPDATE SET
                     material_main=excluded.material_main,
                     sample_count=sample_count+1,
                     updated_at=datetime('now','localtime')""",
                (
                    f"code:{row['material_code']}",
                    row["category"],
                    row["material_code"],
                    row["unit_norm"],
                    city or "",
                    price_tier or "mid",
                    row["material_main"],
                    0.0,
                ),
            )
            upsert_material_fact_from_line(
                conn,
                f"{row['material_code']} {row['material_name']}",
                row["spec_text"],
                row["unit_norm"],
                city or "",
                price_tier or "mid",
                row["material_main"],
                0.0,
            )
        conn.commit()
        total = conn.execute(
            "SELECT COUNT(*) c FROM project_material_prices WHERE project_ref=?",
            (project_ref,),
        ).fetchone()["c"]
    finally:
        conn.close()

    return {
        "imported": len(rows_in),
        "project_ref": project_ref,
        "project_name": project_name,
        "city": city,
        "price_tier": price_tier,
        "catalog_total": total,
        "source": str(fp),
        "codes": [r["material_code"] for r in rows_in],
    }


def collect_st_prices_from_gold(
    gold_path: str | Path,
    *,
    project_ref: str,
    project_name: str,
    city: str = "",
    price_tier: str = "mid",
    db_path: Optional[str] = None,
) -> dict:
    """
    从金标准清单采集 ST 编号→主材价（同编号取中位数）。
    不覆盖已用物料书 xlsx 导入的行；仅补缺失或更新「金标准采集」行。
    """
    from src.knowledge.calibration import load_priced_map

    gold_map = load_priced_map(gold_path)
    by_code: dict[str, list[dict]] = {}
    for row in gold_map.values():
        codes = extract_material_codes(row.name, row.feature or "")
        for code in codes:
            if not code.startswith("ST-"):
                continue
            main = float(row.material_main or 0)
            if main <= 0:
                continue
            by_code.setdefault(code, []).append(
                {
                    "material_main": main,
                    "material_name": (row.name or code).split("\n")[0][:80],
                    "unit_norm": normalize_unit(row.unit or "㎡"),
                }
            )

    conn = get_connection(db_path)
    inserted = updated = skipped = 0
    try:
        for code, samples in sorted(by_code.items()):
            mains = sorted(s["material_main"] for s in samples)
            mid = mains[len(mains) // 2]
            mat_name = samples[0]["material_name"]
            un = samples[0]["unit_norm"] or "㎡"
            existing = conn.execute(
                """SELECT id, source_file, remark FROM project_material_prices
                   WHERE material_code=? AND project_ref=? AND unit_norm=?""",
                (code, project_ref, un),
            ).fetchone()
            if existing and existing["source_file"] and "金标准采集" not in (
                existing["remark"] or ""
            ):
                skipped += 1
                continue
            conn.execute(
                """INSERT INTO project_material_prices
                   (material_code, material_name, name_norm, spec_text, use_area,
                    unit_norm, material_main, category, project_name, project_ref,
                    city, price_tier, source_file, remark)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(material_code, project_ref, unit_norm) DO UPDATE SET
                     material_main=excluded.material_main,
                     material_name=excluded.material_name,
                     name_norm=excluded.name_norm,
                     remark=excluded.remark""",
                (
                    code,
                    mat_name,
                    normalize_name(mat_name),
                    "",
                    "",
                    un,
                    mid,
                    "石材",
                    project_name,
                    project_ref,
                    city or "",
                    price_tier or "mid",
                    str(gold_path),
                    f"金标准采集·{len(samples)}项·中位{mid:.2f}",
                ),
            )
            if existing:
                updated += 1
            else:
                inserted += 1
            conn.execute(
                """INSERT INTO material_price_facts
                   (material_key, material_category, spec_text, unit_norm, city, price_tier,
                    material_main, material_loss_rate, sample_count)
                   VALUES (?,?,?,?,?,?,?,?,1)
                   ON CONFLICT(material_key, unit_norm, city, price_tier) DO UPDATE SET
                     material_main=excluded.material_main,
                     sample_count=sample_count+1,
                     updated_at=datetime('now','localtime')""",
                (
                    f"code:{code}",
                    "石材",
                    mat_name,
                    un,
                    city or "",
                    price_tier or "mid",
                    mid,
                    0.0,
                ),
            )
            upsert_material_fact_from_line(
                conn,
                f"{code} {mat_name}",
                "",
                un,
                city or "",
                price_tier or "mid",
                mid,
                0.0,
            )
        conn.commit()
    finally:
        conn.close()

    return {
        "project_ref": project_ref,
        "codes_seen": len(by_code),
        "inserted": inserted,
        "updated": updated,
        "skipped_xlsx": skipped,
    }


def _is_project_catalog_row(row: dict, project_ref: Optional[str]) -> bool:
    """表价是否来自当前项目价库（非跨项目/城市混用）。"""
    if not project_ref:
        return False
    ref = str(row.get("project_ref") or "")
    return ref == project_ref or ref == f"city:{project_ref}"


def lookup_project_material_row(
    name: str,
    feature: str,
    unit: str,
    *,
    city: str = "",
    price_tier: str = "mid",
    project_ref: Optional[str] = None,
    db_path: Optional[str] = None,
) -> Tuple[Optional[dict], str]:
    """
    按清单名称/特征中的物料编号查项目主材价。
    返回 (row_dict, note) 或 (None, "")。
    """
    codes = extract_material_codes(name, feature)
    if not codes:
        return None, ""

    un = normalize_unit(unit)
    conn = get_connection(db_path)
    try:
        for code in codes:
            for tc in _material_try_codes(code):
                row = None
                from_supplement = False
                if project_ref:
                    row = conn.execute(
                        """SELECT * FROM project_material_prices
                           WHERE material_code=? AND project_ref=?
                             AND (unit_norm=? OR unit_norm='' OR ?='')""",
                        (tc, project_ref, un, un),
                    ).fetchone()
                    if not row:
                        row = _supplement_material_row(
                            tc, project_ref=project_ref, unit=unit, city=city
                        )
                        from_supplement = row is not None
                else:
                    for city_try in ([city, ""] if city else [""]):
                        for tier_try in ([price_tier, "mid"] if price_tier else ["mid"]):
                            row = conn.execute(
                                """SELECT * FROM project_material_prices
                                   WHERE material_code=? AND city=? AND price_tier=?
                                     AND (unit_norm=? OR unit_norm='' OR ?='')
                                   ORDER BY project_ref DESC LIMIT 1""",
                                (tc, city_try, tier_try, un, un),
                            ).fetchone()
                            if row:
                                break
                        if row:
                            break
                    if not row:
                        row = _supplement_material_row(
                            tc, project_ref=None, unit=unit, city=city
                        )
                        from_supplement = row is not None
                if not row:
                    continue
                row_dict = dict(row) if not isinstance(row, dict) else row
                row_dict["_from_project_catalog"] = _is_project_catalog_row(
                    row_dict, project_ref
                )
                stone_name = str(row_dict.get("material_name") or tc)
                prefix = "[补充表价]" if from_supplement else "[项目主材表]"
                note = (
                    f"{prefix}{row_dict.get('project_name', project_ref or '')}·{tc}·"
                    f"{stone_name} 主材{float(row_dict['material_main']):.2f}/"
                    f"{row_dict.get('unit_norm', '')}"
                )
                if tc != code.upper():
                    note = f"{note}（{code}→{tc}）"
                if not row_dict.get("_from_project_catalog") and project_ref:
                    note = f"{note}；⚠非本项目价库，ST/编号表价不跨项目套用"
                return row_dict, note
    finally:
        conn.close()
    return None, ""
